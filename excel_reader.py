from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from config import DASHBOARD_SHEET, REPORT_COLUMNS, SUMMARY_COLUMNS, TAKENAKA_SHEETS, TRACKING_SHEETS
from compare import classify_action, normalize_tracking_state, should_include_tracking
from utils import base_doc_no, contains_doc_no, doc_no_key, extract_category, norm_text, norm_upper, normalize_header, safe_int


def empty_report_df():
    return pd.DataFrame(columns=REPORT_COLUMNS)


def blank_dashboard_matrix():
    return pd.DataFrame([
        {"Status": "Total Document", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0},
        {"Status": "Open", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0},
        {"Status": "on progress", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0},
        {"Status": "Approved", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0},
    ], columns=SUMMARY_COLUMNS)


def row_to_text_and_numbers(ws, row):
    texts = []
    numbers = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        text = norm_text(value)
        if text:
            texts.append(text)
        if isinstance(value, (int, float)):
            numbers.append(safe_int(value))
    return " ".join(texts), numbers


def classify_dashboard_label(row_text):
    text = norm_upper(row_text)
    if "APPROV" in text:
        return "Approved"
    if "PROGRESS" in text or "SEND TO TTI" in text:
        return "on progress"
    if "OPEN" in text or "WAITING ANSWER" in text:
        return "Open"
    if "TOTAL DOCUMENT" in text or ("TOTAL" in text and "DOCUMENT" in text) or "NOT INCLUDED" in text:
        return "Total Document"
    return ""


def parse_dashboard_values(numbers):
    numbers = [safe_int(x) for x in numbers]
    return {
        "Total Document": numbers[0] if len(numbers) > 0 else 0,
        "MAT": numbers[1] if len(numbers) > 1 else 0,
        "MCR": numbers[2] if len(numbers) > 2 else 0,
        "MTS": numbers[3] if len(numbers) > 3 else 0,
        "CVI": numbers[4] if len(numbers) > 4 else 0,
    }


def read_dashboard_sheet(tracking_file):
    wb = load_workbook(tracking_file, data_only=True)
    if DASHBOARD_SHEET not in wb.sheetnames:
        return blank_dashboard_matrix()

    ws = wb[DASHBOARD_SHEET]
    found = {}

    for row in range(1, min(ws.max_row, 50) + 1):
        row_text, numbers = row_to_text_and_numbers(ws, row)
        label = classify_dashboard_label(row_text)
        if not label:
            continue

        values = parse_dashboard_values(numbers)
        if label not in found or values["Total Document"] > found[label]["Total Document"]:
            found[label] = {"Status": label, **values}

    total = found.get("Total Document", {"Status": "Total Document", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0})
    open_row = found.get("Open", {"Status": "Open", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0})
    progress_row = found.get("on progress", {"Status": "on progress", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0})
    approved_row = found.get("Approved", {"Status": "Approved", "Total Document": 0, "MAT": 0, "MCR": 0, "MTS": 0, "CVI": 0})

    for cat in ["MAT", "MCR", "MTS", "CVI"]:
        if total.get(cat, 0) == 0:
            total[cat] = safe_int(open_row.get(cat)) + safe_int(progress_row.get(cat)) + safe_int(approved_row.get(cat))

    return pd.DataFrame([total, open_row, progress_row, approved_row], columns=SUMMARY_COLUMNS)


def get_dashboard_value(matrix, status, column="Total Document"):
    try:
        if matrix is None or matrix.empty or "Status" not in matrix.columns:
            return 0
        matched = matrix.loc[matrix["Status"] == status, column]
        if matched.empty:
            return 0
        return safe_int(matched.iloc[0])
    except Exception:
        return 0


def find_header_row_and_columns(ws):
    best_row = 1
    best_headers = {}
    best_score = -1

    for row in range(1, 30):
        headers = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value:
                headers[normalize_header(value)] = col

        score = 0
        for key in headers:
            if "DOCUMENT" in key:
                score += 3
            if key in ["STATUS", "INFO", "INFORMATION", "VERSION"]:
                score += 3
            if "NAME" in key or "DESCRIPTION" in key:
                score += 2

        if score > best_score:
            best_score = score
            best_row = row
            best_headers = headers

    return best_row, best_headers


def get_col(headers, possible_names):
    for name in possible_names:
        key = normalize_header(name)
        if key in headers:
            return headers[key]
    for key, col in headers.items():
        for name in possible_names:
            name_key = normalize_header(name)
            if name_key in key or key in name_key:
                return col
    return None


def find_doc_no_in_row(ws, row):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if contains_doc_no(value):
            return value, col
    return None, None


def guess_status_from_row(ws, row, status_col=None):
    if status_col:
        value = ws.cell(row=row, column=status_col).value
        if norm_text(value):
            return value
    for col in range(1, ws.max_column + 1):
        value = norm_upper(ws.cell(row=row, column=col).value)
        if value in ["OPEN", "CLOSE", "CLOSED", "RETURNED", "ON PROGRESS", "ON PROCESS", "APPROVED"]:
            return ws.cell(row=row, column=col).value
    return ""


def guess_info_from_row(ws, row, info_col=None):
    if info_col:
        value = ws.cell(row=row, column=info_col).value
        if norm_text(value):
            return value
    for col in range(1, ws.max_column + 1):
        value = norm_upper(ws.cell(row=row, column=col).value)
        if any(word in value for word in ["TTI TO CB", "REVISED", "RESUBMIT", "NA", "ANSWERED"]):
            return ws.cell(row=row, column=col).value
    return ""


def guess_doc_name(ws, row, doc_no_col, doc_name_col=None):
    if doc_name_col:
        value = ws.cell(row=row, column=doc_name_col).value
        if norm_text(value):
            return value
    for col in range(doc_no_col + 1, min(ws.max_column, doc_no_col + 4) + 1):
        value = ws.cell(row=row, column=col).value
        text = norm_text(value)
        if text and "DETH-NSC" not in text:
            return value
    return ""


def read_tracking_all(tracking_file):
    wb = load_workbook(tracking_file, data_only=True)
    rows = []

    for sheet in TRACKING_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        header_row, headers = find_header_row_and_columns(ws)
        doc_no_col = get_col(headers, ["Document No.", "Document No", "Ref No.", "Ref No", "Drawing ref No."])
        doc_name_col = get_col(headers, ["Document Name", "Equipment Name", "Description"])
        category_col = get_col(headers, ["Category", "Document Category"])
        status_col = get_col(headers, ["Status"])
        info_col = get_col(headers, ["Info", "Information"])

        seen = set()

        for row in range(header_row + 1, ws.max_row + 1):
            doc_no = ws.cell(row=row, column=doc_no_col).value if doc_no_col else None
            doc_no_source_col = doc_no_col
            if not doc_no or "DETH-NSC" not in str(doc_no):
                doc_no, doc_no_source_col = find_doc_no_in_row(ws, row)
            if not doc_no or "DETH-NSC" not in str(doc_no):
                continue

            key = doc_no_key(doc_no)
            if (sheet, key) in seen:
                continue
            seen.add((sheet, key))

            category_value = ws.cell(row=row, column=category_col).value if category_col else ""
            category = extract_category(doc_no, category_value)
            doc_name = guess_doc_name(ws, row, doc_no_source_col or 1, doc_name_col)
            tracking_status = guess_status_from_row(ws, row, status_col)
            info = guess_info_from_row(ws, row, info_col)

            rows.append({
                "Tracking Sheet": sheet,
                "Document No": doc_no,
                "Match Key": key,
                "Document Name": doc_name,
                "Document Category": category,
                "Tracking Status": tracking_status,
                "Info": info,
                "Tracking State": normalize_tracking_state(tracking_status, info),
            })

    return pd.DataFrame(rows)


def read_takenaka(takenaka_file):
    wb = load_workbook(takenaka_file, data_only=True)
    data = {}

    for sheet in TAKENAKA_SHEETS:
        if sheet not in wb.sheetnames:
            continue

        ws = wb[sheet]
        for row in range(1, ws.max_row + 1):
            doc_no = ws[f"E{row}"].value
            if not doc_no or "DETH-NSC" not in str(doc_no):
                doc_no, _ = find_doc_no_in_row(ws, row)
            if not doc_no or "DETH-NSC" not in str(doc_no):
                continue

            key = doc_no_key(doc_no)
            record = {
                "Takenaka Sheet": sheet,
                "Takenaka Doc No": doc_no,
                "Takenaka Status 1": ws[f"AA{row}"].value,
                "Takenaka Status 2": ws[f"AB{row}"].value,
                "Takenaka Status 3": ws[f"AC{row}"].value,
            }

            # Keep a record with more status information if duplicated.
            if key not in data:
                data[key] = record
            else:
                existing_score = sum(1 for x in [
                    data[key].get("Takenaka Status 1"),
                    data[key].get("Takenaka Status 2"),
                    data[key].get("Takenaka Status 3"),
                ] if norm_text(x))
                new_score = sum(1 for x in [
                    record.get("Takenaka Status 1"),
                    record.get("Takenaka Status 2"),
                    record.get("Takenaka Status 3"),
                ] if norm_text(x))
                if new_score > existing_score:
                    data[key] = record

    return data


def generate_report(tracking_file, takenaka_file):
    dashboard_matrix = read_dashboard_sheet(tracking_file)
    tracking_df = read_tracking_all(tracking_file)
    takenaka_map = read_takenaka(takenaka_file)

    total_docs = get_dashboard_value(dashboard_matrix, "Total Document")
    open_docs = get_dashboard_value(dashboard_matrix, "Open")
    progress_docs = get_dashboard_value(dashboard_matrix, "on progress")
    approved_docs = get_dashboard_value(dashboard_matrix, "Approved")
    focus_docs = open_docs + progress_docs

    output_wb = load_workbook(tracking_file)
    report_sheet = "Open_On_Process_Compare"
    if report_sheet in output_wb.sheetnames:
        del output_wb[report_sheet]
    report_ws = output_wb.create_sheet(report_sheet)
    report_ws.append(REPORT_COLUMNS)

    fills = {
        "UPDATE TRACKING TO CLOSED": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "OPEN & ON PROCESS": PatternFill(fill_type="solid", fgColor="FFEB9C"),
        "OPEN": PatternFill(fill_type="solid", fgColor="BDD7EE"),
        "OVERDUE / FOLLOW UP": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "RETURNED BY NV5 / NEED RESUBMIT": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "NOT FOUND IN TAKENAKA SOURCE": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "CHECK": PatternFill(fill_type="solid", fgColor="D9EAF7"),
    }

    for cell in report_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    rows = []

    if not tracking_df.empty:
        for _, item in tracking_df.iterrows():
            if not should_include_tracking(item["Tracking Status"], item["Info"]):
                continue

            source = takenaka_map.get(item["Match Key"])
            action, reason = classify_action(source)
            checked_time = datetime.now().strftime("%d-%b-%Y %H:%M:%S")

            if source:
                row_data = [
                    item["Tracking Sheet"], item["Document No"], item["Document Name"], item["Document Category"],
                    item["Tracking Status"], item["Info"],
                    source["Takenaka Sheet"], source["Takenaka Doc No"],
                    source["Takenaka Status 1"], source["Takenaka Status 2"], source["Takenaka Status 3"],
                    action, reason, checked_time
                ]
            else:
                row_data = [
                    item["Tracking Sheet"], item["Document No"], item["Document Name"], item["Document Category"],
                    item["Tracking Status"], item["Info"], "", "", "", "", "", action, reason, checked_time
                ]

            report_ws.append(row_data)
            report_ws[f"L{report_ws.max_row}"].fill = fills.get(action, fills["CHECK"])
            rows.append(dict(zip(REPORT_COLUMNS, row_data)))

    for col in report_ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        report_ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    output = BytesIO()
    output_wb.save(output)
    output.seek(0)

    df = pd.DataFrame(rows, columns=REPORT_COLUMNS)
    action_counts = df["Action"].value_counts().to_dict() if not df.empty else {}

    return {
        "report": output,
        "total_docs": int(total_docs),
        "open_docs": int(open_docs),
        "progress_docs": int(progress_docs),
        "approved_docs": int(approved_docs),
        "focus_docs": int(focus_docs),
        "df": df,
        "tracking_df": tracking_df,
        "dashboard_matrix": dashboard_matrix,
        "action_counts": action_counts,
        "last_updated": datetime.now().strftime("%d-%b-%Y %H:%M:%S"),
    }
